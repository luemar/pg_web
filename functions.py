from datetime import datetime, timedelta
from openpyxl import load_workbook
import pandas as pd
import os, time, sys, traceback
from dateutil.relativedelta import relativedelta
#from config import SHARED_FOLDER, EXCEL_FILE_PATH, DEBUG_ENABLED

SHARED_FOLDER = "/mnt/pg_web_data"
EXCEL_FILE_PATH = os.path.join(SHARED_FOLDER, "PG_2024.xlsx")
DEBUG_ENABLED = True

def debug_log(message):
    if DEBUG_ENABLED:
        timestamp = datetime.now().strftime('%H:%M:%S.%f')[:-3]
        print( f"[DEBUG {timestamp}] {message}")
        sys.stdout.flush()  # Force flush to ensure logs appear immediately

def load_excel_file(force_reload=True):
    debug_log("load_excel_file() function called with force_reload=" + str(force_reload))
    
    global last_file_check, last_modified_time, cached_datetime_dict, cached_string_dict
    try:
        debug_log(f"Attempting to load Excel file from: {EXCEL_FILE_PATH}")

        # Check file existence with detailed information
        if not os.path.exists(EXCEL_FILE_PATH):
            debug_log(f"Excel file NOT FOUND at: {EXCEL_FILE_PATH}")
            debug_log("Contents of shared folder:")
            try:
                files = os.listdir(SHARED_FOLDER)
                for file in files:
                    debug_log(f"  - {file}")
            except Exception as list_error:
                debug_log(f"Error listing directory: {list_error}")

            raise FileNotFoundError(f"Excel file not found at {EXCEL_FILE_PATH}")

        # Check file modification time to detect changes
        current_time = time.time()
        mod_time = os.path.getmtime(EXCEL_FILE_PATH)

        # Only reload if forced or file changed
        if not force_reload and current_time - last_file_check < 5 and mod_time <= last_modified_time:
            # Return cached data if available
            if cached_datetime_dict and cached_string_dict:
                return cached_datetime_dict, cached_string_dict

        # Update check timestamps
        last_file_check = current_time
        last_modified_time = mod_time
		
	# File exists, try to load
        workbook = load_workbook(EXCEL_FILE_PATH, data_only=True)

        # Try to access the 'ages' sheet
        try:
            sheet = workbook["ages"]

            name_list = []
            age_list = []

            # Properly read data from rows
            for row in sheet.iter_rows(min_row=2, max_row=14):
                name = row[0].value
                age = row[2].value
                if name and age:
                    name_list.append(name)
                    age_list.append(age)

            members_dict_datetime = {}
            members_dict_string = {}

            for i, member in enumerate(name_list):
                members_dict_datetime[member] = age_list[i]
                if isinstance(age_list[i], datetime):
                    members_dict_string[member] = age_list[i].strftime('%d.%m.%Y')
                else:
                    debug_log(f"Warning: Age for {member} is not a datetime object: {age_list[i]}")
                    # Handle non-datetime values if needed
                    members_dict_string[member] = str(age_list[i])

            # Cache the results globally
            cached_datetime_dict = members_dict_datetime.copy()
            cached_string_dict = members_dict_string.copy()

            workbook.close()
            debug_log(f"Loaded {len(name_list)} members from Excel")
            return members_dict_datetime, members_dict_string

        except KeyError:
            debug_log("No 'ages' sheet found. Available sheets:")
            debug_log(", ".join(workbook.sheetnames))
            workbook.close()
            raise ValueError("Sheet 'ages' not found in workbook")

    except Exception as e:
        debug_log(f"Error loading Excel file: {e}")
        traceback.print_exc()
        raise

def calculate_average_age():
    debug_log("calculate_average_age() function called")
    
    # Reload Excel data to ensure we have the latest
    try:
        members_dict_datetime, members_dict_string = load_excel_file()
    except Exception as e:
        traceback.print_exc()
        return "Error calculating age"

    ages = []
    today = datetime.now()

    for member, birthday_str in members_dict_string.items():
        try:
            date_format = '%d.%m.%Y'
            birthday = datetime.strptime(birthday_str, date_format)
            age = (today - birthday).days / 365.25
            ages.append(age)
        except Exception as e:
            traceback.print_exc()

    if not ages:
        debug_log("No valid ages found")
        return "No valid ages found"

    average_age = sum(ages) / len(ages)
    debug_log(f"Calculated average age: {average_age:.2f}")
    return '%.2f' % average_age

def calculate_indiv_age(birthdate):
    debug_log(f"calculate_indiv_age() called with birthdate: {birthdate}")
    today = datetime.today()
    years = today.year - birthdate.year
    months = today.month - birthdate.month
    days = today.day - birthdate.day

    if days < 0:
        # Get previous month and year
        previous_month = today.month - 1 if today.month > 1 else 12
        previous_year = today.year if today.month > 1 else today.year - 1
        # Calculate days in previous month
        days_in_previous_month = (datetime(previous_year, previous_month, 1) +
                                relativedelta(months=1) -
                                datetime(previous_year, previous_month, 1)).days

        days += days_in_previous_month
        months -= 1

    if months < 0:
        years -= 1
        months += 12

    debug_log(f"Calculated age: {years} years, {months} months, {days} days")
    return years, months, days

def excel_table():
    debug_log("excel_table() function called")
    
    try:
        # Force reload by resetting cache variables
        global last_file_check, last_modified_time
        last_file_check = 0
        last_modified_time = 0

        # Get the file modification time
        if os.path.exists(EXCEL_FILE_PATH):
            mod_time = os.path.getmtime(EXCEL_FILE_PATH)
        else:
            debug_log(f"WARNING: Excel file not found at: {EXCEL_FILE_PATH}")
            return "Excel file not found"
			
		# IMPORTANT: Use a direct approach to read the Excel file
        try:
            df = pd.read_excel(EXCEL_FILE_PATH, sheet_name="points")
        except Exception as e:
            traceback.print_exc()

            # Try to get available sheet names
            try:
                xls = pd.ExcelFile(EXCEL_FILE_PATH)
                # If "points" sheet doesn't exist, use the first available sheet
                if "points" not in xls.sheet_names and len(xls.sheet_names) > 0:
                    df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=xls.sheet_names[0])
                else:
                    raise Exception("No suitable sheet found")
            except Exception as sheet_error:
                traceback.print_exc()
                return f"Error: Could not find points sheet. {str(e)}"

        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Name'}, inplace=True)

        df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False) | (df.columns == 'Name')]

        for col in df.columns:
            if col != 'Name':
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')
                df[col] = df[col].apply(lambda x: "" if pd.isna(x) else int(x))

        # Add timestamp to class name to force refresh
        timestamp = int(time.time())
        table_html = df.to_html(classes=f'table-container table-{timestamp}', index=False)
        debug_log("Successfully generated HTML table")

        # Return a shorter version for debug purposes
        return table_html
    except Exception as e:
        debug_log(f"Error loading Excel table: {e}")
        traceback.print_exc()
        return f"Error loading Excel table: {str(e)}"
